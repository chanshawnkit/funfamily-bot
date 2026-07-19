import datetime as dt
import os
from typing import Dict, List

import yfinance as yf
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Funfamily_Stock_Tracker.xlsx")
STOCK_SHEET_NAME = "Stock Sheet"

# Excel row numbers (1-based)
HEADER_ROW = 2      # Row 2 contains column headers
DATA_START_ROW = 3  # Row 3 onwards contains data

# Map any ticker aliases to correct Yahoo Finance symbols
TICKER_MAP: Dict[str, str] = {
    "TSMC": "TSM",
}

# Column positions in Excel (1-based)
COL_TICKER         = 8
COL_ORIG_QTY_ANY   = 3
COL_ORIG_QTY_SGD   = 5
COL_ORIG_PRICE_ANY = 10
COL_HOLD_PRICE     = 11
COL_GROSS_VALUE    = 12
COL_NET_PNL        = 13


def get_cell_float(ws, row: int, col: int) -> float:
    """Safely read a cell value as float, returning 0.0 if empty."""
    val = ws.cell(row=row, column=col).value
    try:
        return float(val) if val is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def get_cell_str(ws, row: int, col: int) -> str:
    """Safely read a cell value as string."""
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""


def fetch_price(symbol: str) -> float | None:
    """
    Fetch the latest closing price for a given Yahoo Finance symbol.
    Returns the price as a float, or None if unavailable.
    """
    yf_symbol = TICKER_MAP.get(symbol, symbol)
    print(f"  Fetching {symbol} (as {yf_symbol}) ...")
    try:
        ticker = yf.Ticker(yf_symbol)
        hist = ticker.history(period="5d")

        if hist.empty:
            print(f"  -> No data returned for {yf_symbol}")
            return None

        # Get last available close price as a plain Python float
        raw = hist["Close"].dropna()
        if raw.empty:
            print(f"  -> Close column empty for {yf_symbol}")
            return None

        price = float(raw.values[-1])
        print(f"  -> {symbol}: {price:.4f}")
        return price

    except Exception as e:
        print(f"  -> Error fetching {symbol}: {e}")
        return None


def fetch_quote(symbol: str) -> dict | None:
    """Return the latest quote and day change for any Yahoo Finance ticker."""
    yf_symbol = TICKER_MAP.get(symbol.upper(), symbol.upper())
    try:
        history = yf.Ticker(yf_symbol).history(period="5d")
        if history.empty:
            return None

        closes = history["Close"].dropna()
        if closes.empty:
            return None

        price = float(closes.iloc[-1])
        prev_close = float(closes.iloc[-2]) if len(closes) >= 2 else price
        change = price - prev_close
        change_pct = (change / prev_close * 100) if prev_close else 0.0
        return {
            "symbol": yf_symbol,
            "price": round(price, 4),
            "prev_close": round(prev_close, 4),
            "change": round(change, 4),
            "change_pct": round(change_pct, 2),
        }
    except Exception as exc:
        print(f"  -> Error fetching quote for {symbol}: {exc}")
        return None


def update_prices() -> None:
    """
    Main function:
    1. Open the Excel workbook
    2. Read all stock rows from the Stock Sheet
    3. Fetch latest prices from Yahoo Finance
    4. Update Holding Price, Gross Value and Net P&L cells
    5. Save the workbook
    """
    print(f"Opening: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[STOCK_SHEET_NAME]

    # Step 1 — collect unique tickers and their rows
    ticker_rows: Dict[str, List[int]] = {}
    row = DATA_START_ROW
    while True:
        ticker = get_cell_str(ws, row, COL_TICKER)
        if not ticker:
            break
        if ticker not in ticker_rows:
            ticker_rows[ticker] = []
        ticker_rows[ticker].append(row)
        row += 1

    if not ticker_rows:
        print("No stock rows found in sheet.")
        return

    print(f"Found tickers: {list(ticker_rows.keys())}")

    # Step 2 — fetch prices for all unique tickers
    prices: Dict[str, float] = {}
    for ticker in ticker_rows:
        price = fetch_price(ticker)
        if price is not None:
            prices[ticker] = price

    if not prices:
        print("No prices fetched — nothing to update.")
        return

    # Step 3 — update each row
    updated = 0
    for ticker, excel_rows in ticker_rows.items():
        new_price = prices.get(ticker)
        if new_price is None:
            print(f"  Skipping {ticker} — no price available")
            continue

        for excel_row in excel_rows:
            orig_qty_any   = get_cell_float(ws, excel_row, COL_ORIG_QTY_ANY)
            orig_qty_sgd   = get_cell_float(ws, excel_row, COL_ORIG_QTY_SGD)
            orig_price_any = get_cell_float(ws, excel_row, COL_ORIG_PRICE_ANY)

            if orig_price_any <= 0 or orig_qty_any <= 0 or orig_qty_sgd <= 0:
                print(f"  Row {excel_row}: skipping — missing original data")
                continue

            # Calculate units held and FX rate
            units    = orig_qty_any / orig_price_any
            fx_rate  = orig_qty_sgd / orig_qty_any

            # Recalculate values
            gross_sgd = units * new_price * fx_rate
            net_pnl   = gross_sgd - orig_qty_sgd

            # Write updated values back to Excel
            ws.cell(row=excel_row, column=COL_HOLD_PRICE).value  = round(new_price, 4)
            ws.cell(row=excel_row, column=COL_GROSS_VALUE).value = round(gross_sgd, 6)
            ws.cell(row=excel_row, column=COL_NET_PNL).value     = round(net_pnl, 6)

            print(
                f"  Row {excel_row} ({ticker}): "
                f"price={new_price:.4f}, "
                f"gross={gross_sgd:.2f} SGD, "
                f"P&L={net_pnl:.2f} SGD"
            )
            updated += 1

    # Step 4 — save
    wb.save(EXCEL_PATH)
    print(f"\nUpdated {updated} rows. Saved to {EXCEL_PATH}")


if __name__ == "__main__":
    print(f"[{dt.datetime.now()}] Running price updater...")
    update_prices()
    print(f"[{dt.datetime.now()}] Done.")

