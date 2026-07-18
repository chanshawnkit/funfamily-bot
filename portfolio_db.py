import os
from datetime import date, datetime, timezone
from pathlib import Path

import psycopg
from openpyxl import load_workbook
from psycopg.rows import dict_row

from price_updater import fetch_price


WORKBOOK_PATH = Path(__file__).with_name("Funfamily_Stock_Tracker.xlsx")


def database_url() -> str:
    value = os.getenv("DATABASE_URL") or os.getenv("POSTGRES_URL")
    if not value:
        raise RuntimeError("DATABASE_URL (or POSTGRES_URL) is not configured")
    return value


def connect():
    return psycopg.connect(database_url(), row_factory=dict_row)


def ensure_schema() -> None:
    with connect() as conn, conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS positions (
                id BIGSERIAL PRIMARY KEY,
                purchaser TEXT NOT NULL,
                purchase_date DATE NOT NULL,
                amount_any DOUBLE PRECISION NOT NULL,
                currency TEXT NOT NULL,
                amount_sgd DOUBLE PRECISION NOT NULL,
                platform TEXT NOT NULL,
                product TEXT,
                ticker TEXT NOT NULL,
                reference TEXT,
                original_price_any DOUBLE PRECISION,
                holding_price_any DOUBLE PRECISION,
                gross_value_sgd DOUBLE PRECISION,
                net_pnl_sgd DOUBLE PRECISION,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS positions_member_idx ON positions (LOWER(purchaser))")
        cur.execute("CREATE INDEX IF NOT EXISTS positions_ticker_idx ON positions (UPPER(ticker))")
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS telegram_updates (
                update_id BIGINT PRIMARY KEY,
                received_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute("SELECT pg_advisory_xact_lock(hashtext('funfamily_workbook_seed'))")
        cur.execute("SELECT COUNT(*) AS count FROM positions")
        if cur.fetchone()["count"] == 0:
            _seed_from_workbook(cur)


def _seed_from_workbook(cur) -> None:
    if not WORKBOOK_PATH.exists():
        return
    ws = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)["Stock Sheet"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[7]:
            continue
        purchase_date = row[1]
        if isinstance(purchase_date, datetime):
            purchase_date = purchase_date.date()
        elif not isinstance(purchase_date, date):
            purchase_date = date.fromisoformat(str(purchase_date))
        cur.execute(
            """
            INSERT INTO positions (
                purchaser, purchase_date, amount_any, currency, amount_sgd,
                platform, product, ticker, reference, original_price_any,
                holding_price_any, gross_value_sgd, net_pnl_sgd
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (row[0], purchase_date, row[2], row[3], row[4], row[5], row[6],
             row[7], row[8], row[9], row[10], row[11], row[12]),
        )


def claim_telegram_update(update_id: int) -> bool:
    with connect() as conn, conn.cursor() as cur:
        cur.execute(
            "INSERT INTO telegram_updates (update_id) VALUES (%s) ON CONFLICT DO NOTHING RETURNING update_id",
            (update_id,),
        )
        return cur.fetchone() is not None


def release_telegram_update(update_id: int) -> None:
    """Allow Telegram to retry an update whose processing failed."""
    with connect() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM telegram_updates WHERE update_id = %s", (update_id,))


def positions(member: str | None = None) -> list[dict]:
    query = "SELECT * FROM positions"
    params = ()
    if member:
        query += " WHERE LOWER(purchaser) = LOWER(%s)"
        params = (member,)
    query += " ORDER BY purchaser, purchase_date, id"
    with connect() as conn, conn.cursor() as cur:
        cur.execute(query, params)
        return cur.fetchall()


def add_position(data: dict) -> int:
    with connect() as conn, conn.cursor() as cur:
        price = float(data["price_any"])
        amount_sgd = float(data["amount_sgd"])
        cur.execute(
            """
            INSERT INTO positions (
                purchaser, purchase_date, amount_any, currency, amount_sgd,
                platform, product, ticker, reference, original_price_any,
                holding_price_any, gross_value_sgd, net_pnl_sgd
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
            RETURNING id
            """,
            (data["member"], date.fromisoformat(data["purchase_date"]),
             float(data["amount_any"]), data["currency"].upper(), amount_sgd,
             data["platform"], data.get("product"), data["ticker"].upper(),
             data.get("reference"), price, price, amount_sgd),
        )
        return cur.fetchone()["id"]


def remove_position(position_id: int) -> bool:
    with connect() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM positions WHERE id = %s RETURNING id", (position_id,))
        return cur.fetchone() is not None


def update_ticker_price(ticker: str, price: float) -> int:
    with connect() as conn, conn.cursor() as cur:
        cur.execute(
            """
            UPDATE positions
            SET holding_price_any = %s,
                gross_value_sgd = (amount_any / original_price_any) * %s * (amount_sgd / amount_any),
                net_pnl_sgd = ((amount_any / original_price_any) * %s * (amount_sgd / amount_any)) - amount_sgd,
                updated_at = NOW()
            WHERE UPPER(ticker) = UPPER(%s)
              AND original_price_any > 0 AND amount_any > 0 AND amount_sgd > 0
            """,
            (price, price, price, ticker),
        )
        return cur.rowcount


def refresh_prices() -> tuple[int, list[str]]:
    rows = positions()
    tickers = sorted({row["ticker"] for row in rows})
    updated = 0
    missing = []
    for ticker in tickers:
        price = fetch_price(ticker)
        if price is None:
            missing.append(ticker)
        else:
            updated += update_ticker_price(ticker, price)
    return updated, missing


def fmt_sgd(value: float) -> str:
    return f"S${value:,.2f}"


def fmt_pct(value: float) -> str:
    return f"{'+' if value >= 0 else ''}{value:.2f}%"


def portfolio_summary(member: str | None = None, detailed: bool = False) -> str:
    rows = positions(member)
    if not rows:
        return f"No holdings found for {member}." if member else "No holdings found."
    if detailed and member:
        lines = [f"{member}'s holdings"]
        for row in rows:
            if row["gross_value_sgd"] is None:
                lines.append(f"#{row['id']} {row['ticker']}: needs price data")
                continue
            pct = row["net_pnl_sgd"] / row["amount_sgd"] * 100 if row["amount_sgd"] else 0
            lines.append(
                f"#{row['id']} {row['ticker']} ({row['platform']}): "
                f"{fmt_sgd(row['gross_value_sgd'])}, P&L {fmt_sgd(row['net_pnl_sgd'])} ({fmt_pct(pct)})"
            )
        return "\n".join(lines)

    totals: dict[str, dict[str, float]] = {}
    for row in rows:
        if row["gross_value_sgd"] is None or row["net_pnl_sgd"] is None:
            continue
        item = totals.setdefault(row["purchaser"], {"invested": 0, "value": 0, "pnl": 0})
        item["invested"] += row["amount_sgd"]
        item["value"] += row["gross_value_sgd"]
        item["pnl"] += row["net_pnl_sgd"]
    lines = ["Family portfolio"]
    for name, item in totals.items():
        pct = item["pnl"] / item["invested"] * 100 if item["invested"] else 0
        lines.append(f"{name}: {fmt_sgd(item['value'])}, P&L {fmt_sgd(item['pnl'])} ({fmt_pct(pct)})")
    invested = sum(item["invested"] for item in totals.values())
    value = sum(item["value"] for item in totals.values())
    pnl = sum(item["pnl"] for item in totals.values())
    pct = pnl / invested * 100 if invested else 0
    lines.append(f"Total: {fmt_sgd(value)}, P&L {fmt_sgd(pnl)} ({fmt_pct(pct)})")
    return "\n".join(lines)


def daily_report() -> str:
    report = f"Daily update - {datetime.now(timezone.utc):%d %b %Y}\n\n{portfolio_summary()}"
    missing = sorted({row["ticker"] for row in positions() if row["holding_price_any"] is None})
    return report + (f"\n\nNeeds price data: {', '.join(missing)}" if missing else "")
