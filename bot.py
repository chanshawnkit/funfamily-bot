import asyncio
import logging
import os
from datetime import date, datetime, time
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)
from price_updater import update_prices
import anthropic
from config import validate_anthropic_env

import pandas as pd
from openpyxl import load_workbook as openpyxl_load

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Funfamily_Stock_Tracker.xlsx")
STOCK_SHEET  = "Stock Sheet"
DEPOSIT_SHEET = "Deposit Sheet"

HEADER_ROW = 2  # Excel row containing column headers for Stock Sheet

SELECT_MEMBER, ENTER_TICKER, ENTER_DATE, ENTER_AMOUNT_ANY, ENTER_PRICE, SELECT_CURRENCY, ENTER_AMOUNT_SGD, SELECT_PLATFORM = range(8)
REM_SELECT_MEMBER, REM_SELECT_TICKER, REM_ENTER_DATE, REM_CONFIRM = range(8, 12)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


def load_stock() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_PATH, sheet_name=STOCK_SHEET, header=1)
    df = df.dropna(how="all")
    df = df[df["Ticker"].notna()]
    return df


def append_stock_position(
    purchaser: str,
    purchase_date: date | str,
    amount_any: float,
    currency: str,
    amount_sgd: float,
    platform: str,
    ticker: str,
    original_price_any: float | None = None,
    product: str | None = None,
    reference: str | None = None,
) -> None:
    """
    Append a new stock position as a row in the Stock Sheet without modifying
    existing data or headers.
    """
    wb = openpyxl_load(EXCEL_PATH)
    if STOCK_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{STOCK_SHEET}' not found in workbook")

    ws = wb[STOCK_SHEET]

    # Build a mapping from header name -> column index based on HEADER_ROW.
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=col).value
        if val is None:
            continue
        header_name = str(val).strip()
        if header_name:
            header_map[header_name] = col

    required_headers = [
        "Purchaser",
        "Date of Purchase",
        "Original Purchase Quantum(Any $)",
        "Currency",
        "Original Purchase Quantum(S$)",
        "Platform",
        "Product",
        "Ticker",
        "Reference",
        "Original Purchase Price (Any $)",
        "Holding Price (Any $)",
        "Gross Holding Value (S$)",
        "Net Earning/Loss (S$)",
    ]

    missing = [h for h in required_headers if h not in header_map]
    if missing:
        raise RuntimeError(f"Missing expected columns in Stock Sheet: {', '.join(missing)}")

    # Determine the next empty row after existing data.
    next_row = ws.max_row + 1

    def set_by_header(header: str, value):
        col_idx = header_map.get(header)
        if col_idx is not None:
            ws.cell(row=next_row, column=col_idx).value = value

    # Normalize date to a format consistent with existing rows.
    if isinstance(purchase_date, date):
        date_value = purchase_date
    else:
        date_value = str(purchase_date)

    set_by_header("Purchaser", purchaser)
    set_by_header("Date of Purchase", date_value)
    set_by_header("Original Purchase Quantum(Any $)", float(amount_any))
    set_by_header("Currency", currency)
    set_by_header("Original Purchase Quantum(S$)", float(amount_sgd))
    set_by_header("Platform", platform)
    set_by_header("Product", product or "")
    set_by_header("Ticker", ticker)
    set_by_header("Reference", reference or "")

    # Seed price/valuation columns so units and P&L are well-defined.
    if original_price_any is not None:
        set_by_header("Original Purchase Price (Any $)", float(original_price_any))
        # At purchase time, holding price equals purchase price, value is invested amount, P&L is zero.
        set_by_header("Holding Price (Any $)", float(original_price_any))
        set_by_header("Gross Holding Value (S$)", float(amount_sgd))
        set_by_header("Net Earning/Loss (S$)", 0.0)
    else:
        set_by_header("Original Purchase Price (Any $)", None)
        set_by_header("Holding Price (Any $)", None)
        set_by_header("Gross Holding Value (S$)", None)
        set_by_header("Net Earning/Loss (S$)", None)

    wb.save(EXCEL_PATH)


def fmt_sgd(val: float) -> str:
    return f"S${val:,.2f}"


def fmt_pct(val: float) -> str:
    sign = "+" if val >= 0 else ""
    return f"{sign}{val:.2f}%"


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(help_text())


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(help_text())


async def portfolio_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = load_stock()
    grouped = df.groupby("Purchaser").agg(
        invested=("Original Purchase Quantum(S$)", "sum"),
        value=("Gross Holding Value (S$)", "sum"),
        pnl=("Net Earning/Loss (S$)", "sum"),
    ).reset_index()

    lines = ["*Family Portfolio*\n"]
    total_inv = total_val = total_pnl = 0.0

    for _, r in grouped.iterrows():
        pct = (r.pnl / r.invested * 100) if r.invested else 0
        lines.append(
            f"*{r.Purchaser}*\n"
            f"Invested: {fmt_sgd(r.invested)}\n"
            f"Value:    {fmt_sgd(r.value)}\n"
            f"P&L:      {fmt_sgd(r.pnl)} ({fmt_pct(pct)})\n"
        )
        total_inv += r.invested
        total_val += r.value
        total_pnl += r.pnl

    total_pct = (total_pnl / total_inv * 100) if total_inv else 0
    lines.append(
        f"*Total*\n"
        f"Invested: {fmt_sgd(total_inv)}\n"
        f"Value:    {fmt_sgd(total_val)}\n"
        f"P&L:      {fmt_sgd(total_pnl)} ({fmt_pct(total_pct)})"
    )

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def mystocks_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage: /mystocks <name>\nExample: /mystocks Jasmine")
        return

    member = " ".join(context.args).strip()
    df = load_stock()
    rows = df[df["Purchaser"].str.strip().str.lower() == member.lower()]

    if rows.empty:
        await update.message.reply_text(f"No holdings found for {member}.")
        return

    lines = [f"*{member}'s Holdings*\n"]
    for _, r in rows.iterrows():
        qty   = float(r.get("Original Purchase Quantum(Any $)", 0) or 0)
        price = float(r.get("Original Purchase Price (Any $)", 0) or 0)
        units = qty / price if price > 0 else 0
        hold  = float(r.get("Holding Price (Any $)", 0) or 0)
        val   = float(r.get("Gross Holding Value (S$)", 0) or 0)
        inv   = float(r.get("Original Purchase Quantum(S$)", 0) or 0)
        pnl   = float(r.get("Net Earning/Loss (S$)", 0) or 0)
        pct   = (pnl / inv * 100) if inv else 0

        lines.append(
            f"*{r['Ticker']}* ({r['Platform']})\n"
            f"Units: {units:.3f}\n"
            f"Price: {hold:.4f} {r['Currency']}\n"
            f"Value: {fmt_sgd(val)}\n"
            f"P&L:   {fmt_sgd(pnl)} ({fmt_pct(pct)})\n"
        )

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def pnl_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = load_stock()
    grouped = df.groupby("Purchaser").agg(
        invested=("Original Purchase Quantum(S$)", "sum"),
        pnl=("Net Earning/Loss (S$)", "sum"),
    ).reset_index()

    lines = ["*P&L by Member*\n"]
    for _, r in grouped.iterrows():
        pct = (r.pnl / r.invested * 100) if r.invested else 0
        lines.append(
            f"*{r.Purchaser}*\n"
            f"P&L: {fmt_sgd(r.pnl)} ({fmt_pct(pct)})\n"
        )

    total_pnl = grouped["pnl"].sum()
    total_inv = grouped["invested"].sum()
    total_pct = (total_pnl / total_inv * 100) if total_inv else 0
    lines.append(f"*Total P&L*\n{fmt_sgd(total_pnl)} ({fmt_pct(total_pct)})")

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def update_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) != 2:
        await update.message.reply_text("Usage: /update TICKER PRICE\nExample: /update MSFT 395.50")
        return
    ticker = context.args[0].upper()
    try:
        price = float(context.args[1])
    except ValueError:
        await update.message.reply_text("Price must be a number, e.g. 395.50")
        return

    wb = openpyxl_load(EXCEL_PATH)
    ws = wb[STOCK_SHEET]

    updated = 0
    row = 3
    while True:
        cell_ticker = ws.cell(row=row, column=8).value
        if not cell_ticker:
            break
        if str(cell_ticker).strip().upper() == ticker:
            orig_qty  = float(ws.cell(row=row, column=3).value or 0)
            orig_sgd  = float(ws.cell(row=row, column=5).value or 0)
            orig_price = float(ws.cell(row=row, column=10).value or 0)
            if orig_price > 0 and orig_qty > 0 and orig_sgd > 0:
                units = orig_qty / orig_price
                fx    = orig_sgd / orig_qty
                gross = units * price * fx
                pnl   = gross - orig_sgd
                ws.cell(row=row, column=11).value = round(price, 4)
                ws.cell(row=row, column=12).value = round(gross, 6)
                ws.cell(row=row, column=13).value = round(pnl, 6)
                updated += 1
        row += 1

    if updated == 0:
        await update.message.reply_text(f"Ticker {ticker} not found.")
        return

    wb.save(EXCEL_PATH)
    await update.message.reply_text(
        f"Updated {ticker} to {price:.4f}. {updated} row(s) saved."
    )


async def addstock_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = load_stock()
    members = sorted(df["Purchaser"].dropna().unique().tolist())
    if not members:
        await update.message.reply_text(
            "No existing members found in the stock sheet. Please add at least one row in Excel first."
        )
        return ConversationHandler.END

    keyboard = [[m] for m in members]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    context.user_data["addstock"] = {}
    await update.message.reply_text(
        "Please select a member:", reply_markup=reply_markup
    )
    return SELECT_MEMBER


async def addstock_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    member = (update.message.text or "").strip()
    if not member:
        await update.message.reply_text("Please choose a member from the list.")
        return SELECT_MEMBER

    context.user_data.setdefault("addstock", {})["member"] = member
    await update.message.reply_text(
        "Please enter the ticker symbol (e.g. AAPL, 0P0001Q0TW.SI):",
        reply_markup=ReplyKeyboardRemove(),
    )
    return ENTER_TICKER


async def addstock_ticker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ticker = (update.message.text or "").strip()
    if not ticker:
        await update.message.reply_text("Ticker cannot be empty. Please enter a ticker:")
        return ENTER_TICKER

    context.user_data.setdefault("addstock", {})["ticker"] = ticker
    await update.message.reply_text(
        "Please enter the purchase date in YYYY-MM-DD format (e.g. 2026-03-15):"
    )
    return ENTER_DATE


async def addstock_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_text = (update.message.text or "").strip()
    try:
        purchase_date = date.fromisoformat(date_text)
    except ValueError:
        await update.message.reply_text(
            "Date must be in YYYY-MM-DD format, e.g. 2026-03-15. Please try again:"
        )
        return ENTER_DATE

    context.user_data.setdefault("addstock", {})["purchase_date"] = purchase_date
    await update.message.reply_text(
        "Please enter the original purchase quantum (Any $), e.g. 5000:"
    )
    return ENTER_AMOUNT_ANY


async def addstock_amount_any(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").replace(",", "").strip()
    try:
        amount_any = float(text)
        if amount_any <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "Amount must be a positive number, e.g. 5000 or 5000.50. Please try again:"
        )
        return ENTER_AMOUNT_ANY

    context.user_data.setdefault("addstock", {})["amount_any"] = amount_any

    await update.message.reply_text(
        "Please enter the purchase price per unit (Any $), e.g. 25.50:"
    )
    return ENTER_PRICE


async def addstock_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").replace(",", "").strip()
    try:
        price_any = float(text)
        if price_any <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "Price must be a positive number, e.g. 25.50. Please try again:"
        )
        return ENTER_PRICE

    context.user_data.setdefault("addstock", {})["price_any"] = price_any

    keyboard = [["SGD", "USD"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(
        "Please select the currency:", reply_markup=reply_markup
    )
    return SELECT_CURRENCY


async def addstock_currency(update: Update, context: ContextTypes.DEFAULT_TYPE):
    currency = (update.message.text or "").strip().upper()
    allowed = {"SGD", "USD"}
    if currency not in allowed:
        await update.message.reply_text(
            "Currency must be SGD or USD. Please tap one of the buttons:"
        )
        return SELECT_CURRENCY

    context.user_data.setdefault("addstock", {})["currency"] = currency
    await update.message.reply_text(
        "Please enter the original purchase quantum in SGD, e.g. 5000:",
        reply_markup=ReplyKeyboardRemove(),
    )
    return ENTER_AMOUNT_SGD


async def addstock_amount_sgd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").replace(",", "").strip()
    try:
        amount_sgd = float(text)
        if amount_sgd <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "Amount in SGD must be a positive number, e.g. 5000 or 5000.50. Please try again:"
        )
        return ENTER_AMOUNT_SGD

    context.user_data.setdefault("addstock", {})["amount_sgd"] = amount_sgd

    df = load_stock()
    platforms = sorted(df["Platform"].dropna().unique().tolist())
    if platforms:
        keyboard = [[p] for p in platforms]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        await update.message.reply_text(
            "Please select the platform (or type a new one):", reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            "Please enter the platform (e.g. Endowus, IBKR):",
            reply_markup=ReplyKeyboardRemove(),
        )
    return SELECT_PLATFORM


async def addstock_platform(update: Update, context: ContextTypes.DEFAULT_TYPE):
    platform = (update.message.text or "").strip()
    if not platform:
        await update.message.reply_text(
            "Platform cannot be empty. Please enter or select a platform:"
        )
        return SELECT_PLATFORM

    data = context.user_data.setdefault("addstock", {})
    data["platform"] = platform

    try:
        append_stock_position(
            purchaser=data["member"],
            purchase_date=data["purchase_date"],
            amount_any=data["amount_any"],
            currency=data["currency"],
            amount_sgd=data["amount_sgd"],
            platform=data["platform"],
            ticker=data["ticker"],
            original_price_any=data.get("price_any"),
        )
    except Exception as e:
        logger.exception("Failed to append stock position")
        await update.message.reply_text(
            f"Error adding stock position: {e}", reply_markup=ReplyKeyboardRemove()
        )
        return ConversationHandler.END

    await update.message.reply_text(
        (
            f"Added stock for {data['member']}: {data['ticker']} on "
            f"{data['purchase_date'].isoformat()} for "
            f"{data['amount_any']:.2f} {data['currency']} "
            f"({data['amount_sgd']:.2f} SGD) via {data['platform']}."
        ),
        reply_markup=ReplyKeyboardRemove(),
    )
    context.user_data.pop("addstock", None)
    return ConversationHandler.END


async def addstock_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("addstock", None)
    await update.message.reply_text(
        "Add stock operation cancelled.", reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END


async def removestock_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    df = load_stock()
    members = sorted(df["Purchaser"].dropna().unique().tolist())
    if not members:
        await update.message.reply_text(
            "No existing members found in the stock sheet. Please add at least one row in Excel first."
        )
        return ConversationHandler.END

    keyboard = [[m] for m in members]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    context.user_data["removestock"] = {}
    await update.message.reply_text(
        "Please select the member whose stock you want to remove:",
        reply_markup=reply_markup,
    )
    return REM_SELECT_MEMBER


async def removestock_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    member = (update.message.text or "").strip()
    if not member:
        await update.message.reply_text("Please choose a member from the list.")
        return REM_SELECT_MEMBER

    df = load_stock()
    rows = df[df["Purchaser"].str.strip().str.lower() == member.lower()]
    if rows.empty:
        await update.message.reply_text(
            f"No holdings found for {member}. Nothing to remove."
        )
        return ConversationHandler.END

    tickers = sorted(rows["Ticker"].dropna().unique().tolist())
    if not tickers:
        await update.message.reply_text(
            f"No tickers found for {member}. Nothing to remove."
        )
        return ConversationHandler.END

    keyboard = [[t] for t in tickers]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)

    data = context.user_data.setdefault("removestock", {})
    data["member"] = member

    await update.message.reply_text(
        "Please select the ticker to remove:",
        reply_markup=reply_markup,
    )
    return REM_SELECT_TICKER


async def removestock_ticker(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ticker = (update.message.text or "").strip()
    if not ticker:
        await update.message.reply_text("Please choose a ticker from the list.")
        return REM_SELECT_TICKER

    data = context.user_data.setdefault("removestock", {})
    data["ticker"] = ticker

    await update.message.reply_text(
        "Please enter the purchase date in YYYY-MM-DD format for the row you want to delete:",
        reply_markup=ReplyKeyboardRemove(),
    )
    return REM_ENTER_DATE


async def removestock_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_text = (update.message.text or "").strip()
    try:
        target_date = date.fromisoformat(date_text)
    except ValueError:
        await update.message.reply_text(
            "Date must be in YYYY-MM-DD format, e.g. 2026-03-15. Please try again:"
        )
        return REM_ENTER_DATE

    data = context.user_data.setdefault("removestock", {})
    data["date"] = target_date

    await update.message.reply_text(
        (
            f"You are about to remove all rows for {data['member']}, "
            f"ticker {data['ticker']}, date {target_date.isoformat()}.\n"
            "Reply YES to confirm, or NO to cancel."
        )
    )
    return REM_CONFIRM


async def removestock_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    answer = (update.message.text or "").strip().lower()
    if answer not in {"yes", "y", "no", "n"}:
        await update.message.reply_text(
            "Please reply YES to confirm, or NO to cancel."
        )
        return REM_CONFIRM

    if answer in {"no", "n"}:
        context.user_data.pop("removestock", None)
        await update.message.reply_text("Remove stock operation cancelled.")
        return ConversationHandler.END

    data = context.user_data.get("removestock") or {}
    member = data.get("member")
    ticker = data.get("ticker")
    target_date = data.get("date")
    if not (member and ticker and target_date):
        await update.message.reply_text(
            "Missing information to remove stock. Please try /removestock again."
        )
        context.user_data.pop("removestock", None)
        return ConversationHandler.END

    try:
        wb = openpyxl_load(EXCEL_PATH)
        if STOCK_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{STOCK_SHEET}' not found in workbook")

        ws = wb[STOCK_SHEET]

        # Build header map for key columns
        header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=HEADER_ROW, column=col).value
            if val is None:
                continue
            name = str(val).strip()
            if name:
                header_map[name] = col

        col_purchaser = header_map.get("Purchaser")
        col_ticker = header_map.get("Ticker")
        col_date = header_map.get("Date of Purchase")
        if not (col_purchaser and col_ticker and col_date):
            raise RuntimeError("Required columns (Purchaser, Ticker, Date of Purchase) not found in sheet.")

        rows_to_delete: list[int] = []
        for row in range(HEADER_ROW + 1, ws.max_row + 1):
            cell_purch = ws.cell(row=row, column=col_purchaser).value
            cell_tic = ws.cell(row=row, column=col_ticker).value
            cell_date = ws.cell(row=row, column=col_date).value

            if not cell_purch or not cell_tic or not cell_date:
                continue

            if str(cell_purch).strip().lower() != member.lower():
                continue
            if str(cell_tic).strip() != ticker:
                continue

            match_date = False
            if isinstance(cell_date, datetime):
                match_date = cell_date.date() == target_date
            elif isinstance(cell_date, date):
                match_date = cell_date == target_date
            else:
                match_date = str(cell_date).strip() == target_date.isoformat()

            if match_date:
                rows_to_delete.append(row)

        if not rows_to_delete:
            await update.message.reply_text(
                "No matching rows found for that member, ticker, and date."
            )
            context.user_data.pop("removestock", None)
            return ConversationHandler.END

        # Delete from bottom to top so row indices remain valid
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)

        wb.save(EXCEL_PATH)
    except Exception as e:
        logger.exception("Failed to remove stock rows")
        await update.message.reply_text(f"Error removing stock rows: {e}")
        context.user_data.pop("removestock", None)
        return ConversationHandler.END

    count = len(rows_to_delete)
    await update.message.reply_text(
        f"Removed {count} row(s) for {member}, ticker {ticker}, date {target_date.isoformat()}."
    )
    context.user_data.pop("removestock", None)
    return ConversationHandler.END


async def removestock_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop("removestock", None)
    await update.message.reply_text("Remove stock operation cancelled.")
    return ConversationHandler.END


async def refresh_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Fetching latest prices, please wait...")
    try:
        update_prices()
        await update.message.reply_text("Prices updated successfully.")
    except Exception as e:
        await update.message.reply_text(f"Error updating prices: {e}")


def get_claude_client() -> anthropic.Anthropic:
    api_key, _ = validate_anthropic_env()
    return anthropic.Anthropic(api_key=api_key)

NL_SYSTEM_PROMPT = (
    "You are the assistant for FunFamily Stock Bot, a family investment portfolio "
    "tracker used in a private Telegram group. Family members are Jasmine and "
    "Shawn Mun. Use the provided tools to fetch real data or make changes - never "
    "guess numbers. All monetary values are in SGD unless stated otherwise. Keep "
    "replies short, warm, and easy to read on mobile."
)

CLAUDE_TOOLS = [
    {
        "name": "get_portfolio_summary",
        "description": "Get invested amount, current value, and P&L per member and total, in SGD.",
        "input_schema": {"type": "object", "properties": {}, "required": []},
    },
    {
        "name": "get_member_holdings",
        "description": "Get detailed stock holdings for one family member.",
        "input_schema": {
            "type": "object",
            "properties": {
                "member": {"type": "string", "description": "e.g. Jasmine or Shawn Mun"}
            },
            "required": ["member"],
        },
    },
    {
        "name": "get_pnl",
        "description": "Get profit and loss summary per member and total.",
        "input_schema": {"type": "object", "properties": {}, "required": []},
    },
    {
        "name": "add_stock_position",
        "description": "Add a new stock or fund purchase to the tracker.",
        "input_schema": {
            "type": "object",
            "properties": {
                "member": {"type": "string"},
                "ticker": {"type": "string", "description": "Yahoo Finance ticker"},
                "purchase_date": {"type": "string", "description": "YYYY-MM-DD"},
                "amount_any": {"type": "number", "description": "Amount spent, original currency"},
                "currency": {"type": "string", "description": "SGD or USD"},
                "amount_sgd": {"type": "number", "description": "Amount spent in SGD"},
                "price_any": {"type": "number", "description": "Price per unit, original currency"},
                "platform": {"type": "string", "description": "e.g. Endowus, IBKR"},
            },
            "required": [
                "member", "ticker", "purchase_date", "amount_any",
                "currency", "amount_sgd", "price_any", "platform",
            ],
        },
    },
    {
        "name": "update_stock_price",
        "description": "Manually set the current price for a ticker and recalculate value and P&L.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ticker": {"type": "string"},
                "price": {"type": "number"},
            },
            "required": ["ticker", "price"],
        },
    },
    {
        "name": "refresh_all_prices",
        "description": "Fetch the latest prices for all holdings from Yahoo Finance.",
        "input_schema": {"type": "object", "properties": {}, "required": []},
    },
]


def tool_get_portfolio_summary() -> str:
    df = load_stock()
    df = df[df["Gross Holding Value (S$)"].notna() & df["Net Earning/Loss (S$)"].notna()]
    grouped = df.groupby("Purchaser").agg(
        invested=("Original Purchase Quantum(S$)", "sum"),
        value=("Gross Holding Value (S$)", "sum"),
        pnl=("Net Earning/Loss (S$)", "sum"),
    ).reset_index()

    lines = []
    total_inv = total_val = total_pnl = 0.0
    for _, r in grouped.iterrows():
        pct = (r.pnl / r.invested * 100) if r.invested else 0
        lines.append(
            f"{r.Purchaser}: invested {fmt_sgd(r.invested)}, "
            f"value {fmt_sgd(r.value)}, P&L {fmt_sgd(r.pnl)} ({fmt_pct(pct)})"
        )
        total_inv += r.invested
        total_val += r.value
        total_pnl += r.pnl

    total_pct = (total_pnl / total_inv * 100) if total_inv else 0
    lines.append(
        f"Total: invested {fmt_sgd(total_inv)}, value {fmt_sgd(total_val)}, "
        f"P&L {fmt_sgd(total_pnl)} ({fmt_pct(total_pct)})"
    )
    return "\n".join(lines)


def tool_get_member_holdings(member: str) -> str:
    df = load_stock()
    rows = df[df["Purchaser"].str.strip().str.lower() == member.lower()]
    if rows.empty:
        return f"No holdings found for {member}."

    lines = []
    incomplete = []
    for _, r in rows.iterrows():
        price = float(r.get("Original Purchase Price (Any $)", 0) or 0)
        hold = float(r.get("Holding Price (Any $)", 0) or 0)
        if price <= 0 or pd.isna(r.get("Holding Price (Any $)")):
            incomplete.append(str(r["Ticker"]))
            continue
        qty = float(r.get("Original Purchase Quantum(Any $)", 0) or 0)
        units = qty / price if price > 0 else 0
        val = float(r.get("Gross Holding Value (S$)", 0) or 0)
        inv = float(r.get("Original Purchase Quantum(S$)", 0) or 0)
        pnl = float(r.get("Net Earning/Loss (S$)", 0) or 0)
        pct = (pnl / inv * 100) if inv else 0
        lines.append(
            f"{r['Ticker']} ({r['Platform']}): {units:.3f} units at "
            f"{hold:.4f} {r['Currency']}, value {fmt_sgd(val)}, "
            f"P&L {fmt_sgd(pnl)} ({fmt_pct(pct)})"
        )
    if incomplete:
        lines.append(
            f"\nNote: {', '.join(incomplete)} has missing price data and was "
            f"excluded. Use /update or ask me to set its price to fix this."
        )
    return "\n".join(lines) if lines else f"No complete holdings found for {member}."


def tool_get_pnl() -> str:
    df = load_stock()
    df = df[df["Net Earning/Loss (S$)"].notna()]
    grouped = df.groupby("Purchaser").agg(
        invested=("Original Purchase Quantum(S$)", "sum"),
        pnl=("Net Earning/Loss (S$)", "sum"),
    ).reset_index()

    lines = []
    for _, r in grouped.iterrows():
        pct = (r.pnl / r.invested * 100) if r.invested else 0
        lines.append(f"{r.Purchaser}: P&L {fmt_sgd(r.pnl)} ({fmt_pct(pct)})")

    total_pnl = grouped["pnl"].sum()
    total_inv = grouped["invested"].sum()
    total_pct = (total_pnl / total_inv * 100) if total_inv else 0
    lines.append(f"Total P&L: {fmt_sgd(total_pnl)} ({fmt_pct(total_pct)})")
    return "\n".join(lines)


def tool_add_stock_position(
    member: str, ticker: str, purchase_date: str, amount_any: float,
    currency: str, amount_sgd: float, price_any: float, platform: str,
) -> str:
    parsed_date = date.fromisoformat(purchase_date)
    append_stock_position(
        purchaser=member,
        purchase_date=parsed_date,
        amount_any=float(amount_any),
        currency=currency,
        amount_sgd=float(amount_sgd),
        platform=platform,
        ticker=ticker,
        original_price_any=float(price_any),
    )
    return (
        f"Added {ticker} for {member} on {purchase_date}: "
        f"{amount_any} {currency} ({amount_sgd} SGD) via {platform}."
    )


def tool_update_stock_price(ticker: str, price: float) -> str:
    wb = openpyxl_load(EXCEL_PATH)
    ws = wb[STOCK_SHEET]
    updated = 0
    row = 3
    while True:
        cell_ticker = ws.cell(row=row, column=8).value
        if not cell_ticker:
            break
        if str(cell_ticker).strip().upper() == ticker.upper():
            orig_qty = float(ws.cell(row=row, column=3).value or 0)
            orig_sgd = float(ws.cell(row=row, column=5).value or 0)
            orig_price = float(ws.cell(row=row, column=10).value or 0)
            if orig_price > 0 and orig_qty > 0 and orig_sgd > 0:
                units = orig_qty / orig_price
                fx = orig_sgd / orig_qty
                gross = units * price * fx
                pnl = gross - orig_sgd
                ws.cell(row=row, column=11).value = round(price, 4)
                ws.cell(row=row, column=12).value = round(gross, 6)
                ws.cell(row=row, column=13).value = round(pnl, 6)
                updated += 1
        row += 1

    if updated == 0:
        return f"Ticker {ticker} not found."
    wb.save(EXCEL_PATH)
    return f"Updated {ticker} to {price:.4f}. {updated} row(s) saved."


def tool_refresh_all_prices() -> str:
    update_prices()
    return "Prices refreshed successfully from Yahoo Finance."


def execute_claude_tool(name: str, tool_input: dict) -> str:
    try:
        if name == "get_portfolio_summary":
            return tool_get_portfolio_summary()
        if name == "get_member_holdings":
            return tool_get_member_holdings(tool_input["member"])
        if name == "get_pnl":
            return tool_get_pnl()
        if name == "add_stock_position":
            return tool_add_stock_position(**tool_input)
        if name == "update_stock_price":
            return tool_update_stock_price(tool_input["ticker"], float(tool_input["price"]))
        if name == "refresh_all_prices":
            return tool_refresh_all_prices()
        return f"Unknown tool: {name}"
    except Exception as e:
        logger.exception(f"Tool {name} failed")
        return f"Error running {name}: {e}"


async def handle_natural_language(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_message = update.message.text
    if not user_message:
        return

    messages = [{"role": "user", "content": user_message}]

    try:
        claude_client = get_claude_client()
        _, model = validate_anthropic_env()
        response = await asyncio.to_thread(
            claude_client.messages.create,
            model=model,
            max_tokens=1024,
            system=NL_SYSTEM_PROMPT,
            tools=CLAUDE_TOOLS,
            messages=messages,
        )

        while response.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": response.content})
            tool_results = []
            for block in response.content:
                if block.type == "tool_use":
                    result_text = await asyncio.to_thread(
                        execute_claude_tool, block.name, block.input
                    )
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result_text,
                    })
            messages.append({"role": "user", "content": tool_results})
            response = await asyncio.to_thread(
                claude_client.messages.create,
                model=model,
                max_tokens=1024,
                system=NL_SYSTEM_PROMPT,
                tools=CLAUDE_TOOLS,
                messages=messages,
            )

        final_text = "".join(
            block.text for block in response.content if block.type == "text"
        )
        await update.message.reply_text(final_text or "Done.")

    except Exception as e:
        logger.exception("Natural language handling failed")
        await update.message.reply_text(f"Sorry, I hit an error: {e}")


def build_daily_report() -> str:
    """Build the daily family and member performance message."""
    df = load_stock()
    complete = df[
        df["Gross Holding Value (S$)"].notna()
        & df["Net Earning/Loss (S$)"].notna()
    ]
    if complete.empty:
        return "Daily portfolio update\n\nNo holdings have complete price data yet."

    grouped = complete.groupby("Purchaser").agg(
        invested=("Original Purchase Quantum(S$)", "sum"),
        value=("Gross Holding Value (S$)", "sum"),
        pnl=("Net Earning/Loss (S$)", "sum"),
    )
    lines = [f"Daily portfolio update - {datetime.now(ZoneInfo('Asia/Singapore')):%d %b %Y}", ""]
    for member, row in grouped.iterrows():
        pct = (row["pnl"] / row["invested"] * 100) if row["invested"] else 0
        lines.append(
            f"{member}: {fmt_sgd(row['value'])} | "
            f"P&L {fmt_sgd(row['pnl'])} ({fmt_pct(pct)})"
        )

    invested = grouped["invested"].sum()
    value = grouped["value"].sum()
    pnl = grouped["pnl"].sum()
    pct = (pnl / invested * 100) if invested else 0
    lines.extend(["", f"Family total: {fmt_sgd(value)}", f"Total P&L: {fmt_sgd(pnl)} ({fmt_pct(pct)})"])

    missing = df[df["Holding Price (Any $)"].isna()]["Ticker"].dropna().astype(str).unique()
    if len(missing):
        lines.extend(["", f"Needs price data: {', '.join(missing)}"])
    return "\n".join(lines)


async def daily_update_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    chat_id = context.job.chat_id
    try:
        await asyncio.to_thread(update_prices)
        await context.bot.send_message(chat_id=chat_id, text=build_daily_report())
    except Exception:
        logger.exception("Daily portfolio update failed for chat %s", chat_id)
        await context.bot.send_message(
            chat_id=chat_id,
            text="The daily portfolio update could not be completed. Check the bot logs.",
        )


def schedule_daily_updates(app: Application) -> int:
    raw_chat_ids = os.getenv("TELEGRAM_DAILY_CHAT_IDS", "")
    chat_ids = [int(value.strip()) for value in raw_chat_ids.split(",") if value.strip()]
    if not chat_ids:
        logger.warning("TELEGRAM_DAILY_CHAT_IDS is empty; daily updates are disabled")
        return 0

    raw_time = os.getenv("DAILY_UPDATE_TIME", "08:00")
    try:
        hour, minute = (int(part) for part in raw_time.split(":", 1))
        scheduled_time = time(hour=hour, minute=minute, tzinfo=ZoneInfo("Asia/Singapore"))
    except (TypeError, ValueError) as exc:
        raise RuntimeError("DAILY_UPDATE_TIME must use 24-hour HH:MM format") from exc

    if app.job_queue is None:
        raise RuntimeError('Install python-telegram-bot with the "job-queue" extra')
    for chat_id in chat_ids:
        app.job_queue.run_daily(
            daily_update_job,
            time=scheduled_time,
            chat_id=chat_id,
            name=f"daily-portfolio-{chat_id}",
        )
    return len(chat_ids)


def help_text() -> str:
    return (
        "/portfolio - Full family portfolio summary\n"
        "/mystocks <name> - One member's holdings\n"
        "/pnl - Profit & loss per member\n"
        "/update <ticker> <price> - Manually set a price\n"
        "/refresh - Fetch latest prices from Yahoo Finance\n"
        "/addstock - Guided flow to add a new stock position\n"
        "/removestock - Guided flow to remove an existing stock row\n"
        "/help - Show this message"
    )


def main():
    load_dotenv()
    validate_anthropic_env()
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN not set in .env")

    app = Application.builder().token(token).build()
    scheduled_count = schedule_daily_updates(app)

    addstock_conv = ConversationHandler(
        entry_points=[CommandHandler("addstock", addstock_start)],
        states={
            SELECT_MEMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_member)],
            ENTER_TICKER: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_ticker)],
            ENTER_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_date)],
            ENTER_AMOUNT_ANY: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_amount_any)],
            ENTER_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_price)],
            SELECT_CURRENCY: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_currency)],
            ENTER_AMOUNT_SGD: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_amount_sgd)],
            SELECT_PLATFORM: [MessageHandler(filters.TEXT & ~filters.COMMAND, addstock_platform)],
        },
        fallbacks=[CommandHandler("cancel", addstock_cancel)],
    )

    removestock_conv = ConversationHandler(
        entry_points=[CommandHandler("removestock", removestock_start)],
        states={
            REM_SELECT_MEMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, removestock_member)],
            REM_SELECT_TICKER: [MessageHandler(filters.TEXT & ~filters.COMMAND, removestock_ticker)],
            REM_ENTER_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, removestock_date)],
            REM_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, removestock_confirm)],
        },
        fallbacks=[CommandHandler("cancel", removestock_cancel)],
    )

    app.add_handler(CommandHandler("start",     start_command))
    app.add_handler(CommandHandler("help",      help_command))
    app.add_handler(CommandHandler("portfolio", portfolio_command))
    app.add_handler(CommandHandler("mystocks",  mystocks_command))
    app.add_handler(CommandHandler("pnl",       pnl_command))
    app.add_handler(CommandHandler("update",    update_command))
    app.add_handler(CommandHandler("refresh",   refresh_command))
    app.add_handler(addstock_conv)
    app.add_handler(removestock_conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_natural_language))

    logger.info("Bot started with %d daily update destination(s).", scheduled_count)
    app.run_polling()


if __name__ == "__main__":
    main()
